"""parity/harness - Excel<->Python parity (Method Spec 10).

Reads sheet '10 Output Contract' from the frozen template and compares the Python
tidy output cell-for-cell. Reports the distribution of relative differences and
fails the build above the assumptions-book tolerance (0.001%). Jess's sheets are
read-only and never modified programmatically; the build only ever adds sheets.
Author: Avia Solutions.
"""
from __future__ import annotations
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from ..config import get

PARITY_SHEET = "10 Output Contract"
_HEADER_ROW = 4          # year headers
_FIRST_METRIC_ROW = 5
_LAST_METRIC_ROW = 23
_FIRST_YEAR_COL = 3      # column C


def read_output_contract(path: str | Path, sheet: str = PARITY_SHEET) -> pd.DataFrame:
    """Extract the tidy frame [metric, dest_region, year, value] from cached values."""
    wb = load_workbook(str(path), data_only=True, read_only=True)
    ws = wb[sheet]
    grid = list(ws.iter_rows(min_row=1, max_row=_LAST_METRIC_ROW, values_only=True))
    header = grid[_HEADER_ROW - 1]
    years = {c: header[c] for c in range(_FIRST_YEAR_COL - 1, len(header))
             if isinstance(header[c], (int, float))}
    rows = []
    for r in range(_FIRST_METRIC_ROW - 1, _LAST_METRIC_ROW):
        line = grid[r]
        metric = line[0]
        if metric is None:
            continue
        region = line[1] if line[1] not in (None, "-") else None
        for c, yr in years.items():
            v = line[c]
            if v is None:
                continue
            rows.append({"metric": metric, "dest_region": region,
                         "year": int(yr), "value": float(v)})
    return pd.DataFrame(rows)


def compare(excel: pd.DataFrame, python: pd.DataFrame, tol: float | None = None) -> pd.DataFrame:
    """Join on (metric, dest_region, year); return a report with relative
    differences and a per-row pass flag. Relative difference falls back to
    absolute when the Excel value is zero."""
    tol = get("reconciliation.parity_tolerance_rel") if tol is None else tol
    keys = ["metric", "dest_region", "year"]
    m = excel.merge(python, on=keys, how="outer", suffixes=("_xl", "_py"), indicator=True)
    denom = m["value_xl"].abs().where(m["value_xl"].abs() > 0, other=np.nan)
    m["abs_diff"] = (m["value_xl"] - m["value_py"]).abs()
    m["rel_diff"] = m["abs_diff"] / denom
    m["rel_diff"] = m["rel_diff"].fillna(m["abs_diff"])   # zero-denominator fallback
    m["within_tol"] = (m["_merge"] == "both") & (m["rel_diff"] <= tol)
    return m


def parity_passes(report: pd.DataFrame) -> bool:
    return bool(report["within_tol"].all())


def summarise(report: pd.DataFrame) -> dict:
    matched = report[report["_merge"] == "both"]
    return {
        "cells": int(len(report)),
        "matched": int(len(matched)),
        "excel_only": int((report["_merge"] == "left_only").sum()),
        "python_only": int((report["_merge"] == "right_only").sum()),
        "max_rel_diff": float(matched["rel_diff"].max()) if len(matched) else 0.0,
        "failing": int((~report["within_tol"]).sum()),
        "passes": parity_passes(report),
    }
