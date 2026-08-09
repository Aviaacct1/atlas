"""outputs/chart_writer - write the impact table and Avia-format charts to xlsx
(Cockpit build update F). Dedicated chart sheet, Office 2024 palette, legend at the
bottom, no gridlines, "Source:" line, author-stamped "Avia Solutions".
Author: Avia Solutions.
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment

from . import chart_format as cf


def add_forecast_chart(wb, chart_data, base_year, heading,
                       data_sheet="_chartdata", chart_sheet="Chart - Forecast"):
    """Add the Avia-format chart to an existing workbook, and return the chartsheet.

    Split out of write_forecast_workbook on 9 August 2026 so the configured-airport
    deliverable can carry the same chart. Until then this module was imported by nothing,
    so every Excel deliverable Atlas produced went out with no chart at all while the
    formatting rules beside it were tested and passing.

    Avia format: dedicated chart sheet, Office 2024 pinned palette, legend at the bottom,
    no gridlines. The chart must stand on its own, so the heading carries what it shows.
    """
    ds = wb.create_sheet(data_sheet)
    years = chart_data["years"]
    names = list(chart_data["series"])
    ds.append(["Year"] + names)
    for y in years:
        ds.append([cf.year_label(y, base_year)] + [chart_data["series"][n][y] for n in names])
    ds.sheet_state = "hidden"          # working data, not a page of the deliverable

    chart = LineChart()
    chart.title = heading
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None
    chart.legend.position = "b"
    data = Reference(ds, min_col=2, max_col=1 + len(names), min_row=1, max_row=1 + len(years))
    chart.add_data(data, titles_from_data=True)
    cats = Reference(ds, min_col=1, min_row=2, max_row=1 + len(years))
    chart.set_categories(cats)
    for i, s in enumerate(chart.series):
        s.graphicalProperties.line.solidFill = cf.PINNED_PALETTE[i % len(cf.PINNED_PALETTE)]
        s.graphicalProperties.line.width = 28000
    cs = wb.create_chartsheet(chart_sheet)
    cs.add_chart(chart)
    return cs


def write_forecast_workbook(path, impact_table, chart_data, base_year, heading,
                            source="OAG, AviaSolutions analysis"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Impact Table"
    sy = impact_table["spot_years"]
    periods = impact_table["cagr_periods"]

    ws["A1"] = heading
    ws["A1"].font = Font(name=cf.FONT, bold=True, size=14)
    header = ["Metric"] + [cf.year_label(y, base_year) for y in sy] + [f"CAGR {a}-{b}" for a, b in periods]
    ws.append([])
    ws.append(header)
    for c in range(1, len(header) + 1):
        ws.cell(3, c).font = Font(name=cf.FONT, bold=True)
    for m, e in impact_table["metrics"].items():
        ws.append([m] + [e["spot"].get(y) for y in sy] +
                  [(round(e["cagr"][f"{a}-{b}"], 4) if e["cagr"].get(f"{a}-{b}") is not None else None)
                   for a, b in periods])
        if "vs_baseline" in e:
            ws.append([f"  {m} vs baseline"] + [round(e["vs_baseline"].get(y, 0.0), 3) for y in sy])
    ws.append([])
    src = ws.cell(ws.max_row + 1, 1, cf.source_line(source))
    src.font = Font(name=cf.FONT, size=11)

    # ---- dedicated chart sheet, Avia format ----
    add_forecast_chart(wb, chart_data, base_year, heading)

    wb.properties.creator = cf.AUTHOR
    wb.properties.lastModifiedBy = cf.AUTHOR
    wb.properties.title = heading
    wb.save(path)
    return path
