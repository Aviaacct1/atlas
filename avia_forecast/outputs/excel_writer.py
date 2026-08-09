"""outputs/excel_writer - generic, config-driven instance deliverable writer (O-10). Author: Avia Solutions.

Builds the client-shaped Excel deliverable from instance.output_rows alone, with no airport-specific
logic. The workbook is normalised (fixed document timestamps and zip member dates) so regeneration is
byte-identical.
"""
from __future__ import annotations
import os
import zipfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font

from ..airports import instance
from . import chart_writer

_LABELS = [("total", "Total passengers"), ("international", "International"),
           ("domestic", "Domestic"), ("transit", "Transit"), ("charter", "Charter"),
           ("ga", "GA / other"), ("non_lcc", "Non-LCC passengers"), ("lcc", "LCC passengers"),
           ("commercial_atm", "Commercial ATMs"), ("landed_tonnage", "Landed tonnage")]


def _normalise_zip(path: str):
    """Rewrite the xlsx zip with sorted members and a fixed date, so the bytes are deterministic."""
    tmp = path + ".norm"
    with zipfile.ZipFile(path) as zin:
        names = sorted(zin.namelist())
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for n in names:
                data = zin.read(n)
                zi = zipfile.ZipInfo(n, date_time=(1980, 1, 1, 0, 0, 0))
                zi.compress_type = zipfile.ZIP_DEFLATED
                zout.writestr(zi, data)
    os.replace(tmp, path)


def write_instance_excel(cfg: dict, path: str, overrides=None) -> str:
    o = instance.output_rows(cfg, overrides)
    spots = o["spots"]
    cwins = list(o["cagr"].keys())
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Forecast"
    name = cfg["meta"].get("airport", "Airport")
    ws["B2"] = f"{name} - Traffic forecast (Avia global forecast engine, configured instance)"
    ws["B2"].font = Font(bold=True, size=14)
    ws["B3"] = "Source: Sabre GDD, OEF/IMF, OAG, AviaSolutions analysis"

    r0 = 5
    ws.cell(row=r0, column=2, value="Series").font = Font(bold=True)
    for j, y in enumerate(spots):
        ws.cell(row=r0, column=3 + j, value=y).font = Font(bold=True)
    for j, w in enumerate(cwins):
        ws.cell(row=r0, column=3 + len(spots) + j, value=f"CAGR {w}").font = Font(bold=True)
    rr = r0 + 1
    for key, lab in _LABELS:
        ws.cell(row=rr, column=2, value=lab)
        for j, y in enumerate(spots):
            v = o["series"].get(key, {}).get(y)
            ws.cell(row=rr, column=3 + j, value=(round(v, 1) if isinstance(v, (int, float)) else v))
        for j, w in enumerate(cwins):
            c = o["cagr"][w].get(key)
            cell = ws.cell(row=rr, column=3 + len(spots) + j, value=(round(c * 100, 2) if c is not None else None))
            if c is not None:
                cell.number_format = '0.00"%"'
        rr += 1
    ws.column_dimensions["B"].width = 26

    # ---- Avia-format chart, on its own sheet ----
    # Added 9 August 2026. outputs/chart_writer.py was imported by nothing, so every
    # configured-airport workbook went to a client with no chart while the formatting
    # rules beside it were tested and green. The heading carries the airport, the unit
    # and the period, so the chart stands on its own away from the workbook.
    by = int(cfg["meta"].get("base_year", 2025))
    yrs = [y for y in o["years"] if isinstance(y, int)]
    chart_series = {}
    for key, lab in (("total", "Total passengers"), ("international", "International"),
                     ("domestic", "Domestic")):
        s = o["series"].get(key) or {}
        vals = {y: s.get(y) for y in yrs}
        if any(v is not None for v in vals.values()):
            chart_series[lab] = {y: (v if isinstance(v, (int, float)) else None) for y, v in vals.items()}
    if chart_series and yrs:
        chart_writer.add_forecast_chart(
            wb, {"years": yrs, "series": chart_series}, by,
            f"{name} passengers, million a year, {yrs[0]} to {yrs[-1]} (forecast from {by})")

    aw = wb.create_sheet("Assumptions")
    aw["A1"] = "Assumptions register"; aw["A1"].font = Font(bold=True)
    aw.append(["group", "input", "resolved", "source", "reason"])
    for row in o.get("assumptions_register", {}).get("rows", []):
        aw.append([row.get("group"), row.get("input"), str(row.get("resolved")),
                   row.get("source"), row.get("reason")])

    wb.properties.creator = "Avia Solutions"
    wb.properties.lastModifiedBy = "Avia Solutions"
    fixed = datetime(int(cfg["meta"].get("base_year", 2025)), 1, 1)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.save(path)
    _normalise_zip(path)
    return path
