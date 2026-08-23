"""O-10 writer half: a generic, config-driven Excel writer built from instance.output_rows (no
007-specific logic), author-stamped, regenerating byte-identically and reconciling to the engine.
Author: Avia Solutions."""
import hashlib
import openpyxl
from avia_forecast.airports import instance, qa
from avia_forecast.outputs import excel_writer


def _sha(p):
    return hashlib.sha256(open(p, "rb").read()).hexdigest()


def _zip_diff(p1, p2):
    """Name the differing members, because two hashes name nothing. Added 23 August 2026
    after this test failed on a machine whose shared system Python held a different
    openpyxl than the venv: the failure message must say WHERE the bytes differ, so an
    environment fault can be told from a writer fault in one read. Run the suite with
    .venv\\Scripts\\python.exe, never the shared interpreter."""
    import zipfile
    z1, z2 = zipfile.ZipFile(p1), zipfile.ZipFile(p2)
    out = []
    for n in sorted(set(z1.namelist()) | set(z2.namelist())):
        a = z1.read(n) if n in z1.namelist() else b"<absent>"
        b = z2.read(n) if n in z2.namelist() else b"<absent>"
        if a != b:
            i = next((k for k, (x, y) in enumerate(zip(a, b)) if x != y), min(len(a), len(b)))
            out.append(f"{n}: first difference at byte {i} ({len(a)} vs {len(b)} bytes)")
    return out


def test_generic_writer_deterministic_and_stamped(tmp_path):
    cfg = instance.load("zagreb")
    p1 = tmp_path / "z1.xlsx"; p2 = tmp_path / "z2.xlsx"
    excel_writer.write_instance_excel(cfg, str(p1))
    excel_writer.write_instance_excel(cfg, str(p2))
    assert _sha(str(p1)) == _sha(str(p2)), (                    # byte-comparable regeneration
        "two consecutive writes differ; differing members: "
        + "; ".join(_zip_diff(str(p1), str(p2))))
    wb = openpyxl.load_workbook(str(p1))
    assert wb.properties.creator == "Avia Solutions" and "Assumptions" in wb.sheetnames
    assert qa.check_author_stamp(str(p1))["ok"]


def test_generic_writer_reconciles_to_engine(tmp_path):
    cfg = instance.load("zagreb")
    o = instance.output_rows(cfg)
    p = tmp_path / "z.xlsx"; excel_writer.write_instance_excel(cfg, str(p))
    ws = openpyxl.load_workbook(str(p))["Forecast"]
    col = 3 + o["spots"].index(2045)
    trow = next(r for r in range(1, 60) if ws.cell(row=r, column=2).value == "Total passengers")
    val = ws.cell(row=trow, column=col).value
    assert abs(val - round(o["series"]["total"][2045], 1)) < 1.0


def test_generic_writer_runs_for_arbitrary_config(tmp_path):
    cfg = {"meta": {"airport": "TST", "base_year": 2025, "horizon": 2030},
           "markets": {"m": {"elasticity": 1.0, "base_weight": 1.0}},
           "gdp_index": {"m": {str(y): 1.0 + 0.02 * (y - 2025) for y in range(2025, 2031)}},
           "base": {"total": 1000.0, "international": 1000.0, "domestic": 0, "transit": 0,
                    "non_lcc": 1000.0, "lcc": 0, "commercial_atm": 10.0}}
    p = tmp_path / "t.xlsx"
    excel_writer.write_instance_excel(cfg, str(p))
    assert qa.check_author_stamp(str(p))["ok"]


def test_workbook_carries_an_avia_format_chart(tmp_path):
    """The deliverable must contain a chart, not merely a module that could draw one.

    outputs/chart_writer.py was imported by nothing until 9 August 2026, so every
    configured-airport workbook went out with no chart while tests/test_impact_and_charts.py
    passed on the formatting rules beside it. That is the shape the capability audit exists
    to catch: a green suite says nothing about what it was never asked. This asks.
    """
    import zipfile
    cfg = instance.load("zagreb")
    p = tmp_path / "chart.xlsx"
    excel_writer.write_instance_excel(cfg, str(p))
    wb = openpyxl.load_workbook(str(p))
    charts = [s for s in wb.sheetnames if s.startswith("Chart -")]
    assert charts, f"no chart sheet in the workbook; sheets are {wb.sheetnames}"
    with zipfile.ZipFile(str(p)) as z:
        parts = z.namelist()
    assert any(n.startswith("xl/charts/chart") for n in parts), (
        "a chart sheet exists but the file holds no chart part")
    assert wb.properties.creator == "Avia Solutions"
    assert wb.properties.lastModifiedBy == "Avia Solutions"
