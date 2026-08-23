"""Ingest an IMF WEO 'all countries' xlsx: extract each country's real GDP growth (subject code
NGDP_RPCH) and cumulate to a forecast GDP index {ISO2: {year: index}} for the backtest decomposition.
IMF is ISO3; mapped to ISO2 (pycountry, with a small fallback). Author: Avia Solutions.

Usage: python scripts/ingest_imf_weo.py <weo.xlsx> <base_year> [out.json]
"""
import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from avia_forecast.io_safe import dump_atomic  # atomic write with parse-back (io_safe rule, 23 Aug 2026)

try:
    import pycountry
    def iso3to2(c):
        try:
            return pycountry.countries.get(alpha_3=c).alpha_2
        except Exception:
            return None
except Exception:
    _FB = {"USA": "US", "GBR": "GB", "DEU": "DE", "FRA": "FR", "ESP": "ES", "ITA": "IT",
           "CHN": "CN", "IND": "IN", "TUR": "TR", "NLD": "NL", "RUS": "RU", "JPN": "JP"}
    def iso3to2(c):
        return _FB.get(c)


def extract_growth(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    ycols = {}
    for j, v in enumerate(hdr):
        try:
            y = int(str(v).strip())
        except (ValueError, TypeError):
            continue
        if 1980 <= y <= 2060:
            ycols[j] = y
    growth = {}
    for r in rows[1:]:
        if not (len(r) > 3 and r[2] == "NGDP_RPCH"):
            continue
        iso2 = iso3to2(str(r[1]).strip()) if r[1] else None
        if not iso2:
            continue
        series = {}
        for j, y in ycols.items():
            v = r[j]
            if isinstance(v, (int, float)):
                series[y] = float(v)
        if series:
            growth[iso2] = series
    return growth


def to_index(growth, base_year):
    out = {}
    for c, g in growth.items():
        yrs = sorted(y for y in g if y >= base_year)
        if base_year not in g:
            continue
        idx, lvl = {str(base_year): 1.0}, 1.0
        for y in range(base_year + 1, max(yrs) + 1 if yrs else base_year):
            lvl *= (1.0 + g.get(y, 0.0) / 100.0)
            idx[str(y)] = round(lvl, 6)
        out[c] = idx
    return out


def main():
    path, base_year = sys.argv[1], int(sys.argv[2])
    idx = to_index(extract_growth(path), base_year)
    out = sys.argv[3] if len(sys.argv) > 3 else os.path.splitext(path)[0] + f"_gdp_index_{base_year}.json"
    dump_atomic({"base_year": base_year, "gdp_index": idx}, out, indent=0)
    print(f"{len(idx)} countries -> {out}  (base {base_year})")


if __name__ == "__main__":
    main()
