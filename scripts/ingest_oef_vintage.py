"""Ingest an OEF 'GDP Forecasts' vintage xlsx: extract each country's central-case Real GDP Growth (%)
series and cumulate to a forecast GDP index {country_name: {year: index}} for the backtest
decomposition. OEF growth is decimal (0.041 = 4.1%); countries are by name (needs a name->ISO2 join
for the ACI backtest). Author: Avia Solutions.

Usage: python scripts/ingest_oef_vintage.py <vintage.xlsx> <base_year> [out.json]
"""
import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from avia_forecast.io_safe import dump_atomic  # atomic write with parse-back (io_safe rule, 23 Aug 2026)


def _rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return list(wb[wb.sheetnames[0]].iter_rows(values_only=True))


def _year(v):
    try:
        iv = int(str(v).strip())
    except (ValueError, TypeError):
        return None
    return iv if 2000 <= iv <= 2060 else None


def extract_growth(rows):
    """{country_name: {year: real GDP growth fraction}} using the leftmost strictly-increasing year
    block as the central case."""
    ycols = {}
    for r in rows:
        pairs = [(j, _year(v)) for j, v in enumerate(r) if _year(v) is not None]
        if len(pairs) >= 5:
            block = {}
            last = None
            for j, y in pairs:                       # take the first strictly-increasing run (central case)
                if last is None or y > last:
                    block[j] = y; last = y
                else:
                    break
            ycols = block
            break
    if not ycols:
        raise SystemExit("no year header row found")
    growth = {}
    for r in rows:
        name = r[0] if (isinstance(r[0], str) and r[0].strip() and not r[0].strip().endswith(":")) else None
        if not name:
            continue
        series = {ycols[j]: float(r[j]) for j in ycols if isinstance(r[j], (int, float))}
        if len(series) >= 3:
            growth[name.strip()] = series
    return growth


def to_index(growth, base_year):
    out = {}
    for name, g in growth.items():
        yrs = sorted(y for y in g if y >= base_year)
        if base_year not in g:
            continue
        idx, lvl = {base_year: 1.0}, 1.0
        for y in range(base_year + 1, max(yrs) + 1 if yrs else base_year):
            lvl *= (1.0 + g.get(y, 0.0))
            idx[y] = round(lvl, 6)
        out[name] = idx
    return out


def main():
    path, base_year = sys.argv[1], int(sys.argv[2])
    growth = extract_growth(_rows(path))
    idx = to_index(growth, base_year)
    out = sys.argv[3] if len(sys.argv) > 3 else os.path.splitext(path)[0] + f"_gdp_index_{base_year}.json"
    dump_atomic({"base_year": base_year, "gdp_index": idx}, out, indent=0)
    print(f"{len(idx)} countries -> {out}  (base {base_year})")


if __name__ == "__main__":
    main()
