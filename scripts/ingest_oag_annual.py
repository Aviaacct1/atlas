"""ingest_oag_annual - load Jess's full-year OAG Schedules Power Table exports.

Author: Avia Solutions.

Reads the region-period .xlsx files (Europe/Asia monthly; Africa/Middle East/
Latin America half-yearly; North America/Southwest Pacific annual) from
E:\\Avia\\Global\\OAG\\raw, streams each with openpyxl read-only, matches columns
BY NAME, writes one parquet per workbook and appends to a duckdb store.

Resumable: a manifest records every completed file (name + size); re-running
skips completed work, so it is safe to launch in the background while the
Egnyte sync is still filling the raw folder - unfinished downloads are retried
on the next pass.

Usage:
  python ingest_oag_annual.py                 # ingest everything new in raw/
  python ingest_oag_annual.py --one "Europe Jan 2015.xlsx"   # validate one file
  python ingest_oag_annual.py --loop          # keep polling raw/ every 10 min
"""
from __future__ import annotations
import glob, json, os, re, sys, time

import duckdb
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from avia_forecast import paths

# The OAG raw folder is inside the Global root, so it follows AVIA_GLOBAL_ROOT and
# needs no name of its own. AVIA_OAG was a sixth variable for a folder already known.
BASE = os.path.join(paths.GLOBAL, "OAG")
RAW, PARQ = os.path.join(BASE, "raw"), os.path.join(BASE, "parquet")
DB = os.path.join(BASE, "oag_annual.duckdb")
MANIFEST = os.path.join(BASE, "ingest_manifest.json")

WANTED = ["Carrier Code", "Carrier Name", "Dep Airport Code", "Dep City Code",
          "Dep IATA Country Code", "Dep IATA Country Name", "Dep Region Name",
          "Arr Airport Code", "Arr City Code", "Arr IATA Country Code",
          "Arr IATA Country Name", "Arr Region Name", "International/Domestic",
          "Specific Aircraft Code", "Specific Aircraft Name", "GCD (km)",
          "Mainline/Low Cost", "Service Type", "Restrictions", "Frequency",
          "Seats (Total)", "First seats (Total)", "Business seats (Total)",
          "Economy seats (Total)", "Time series", "Year"]
NUMERIC = {"Frequency", "Seats (Total)", "First seats (Total)",
           "Business seats (Total)", "Economy seats (Total)", "GCD (km)"}
KEY = {"Carrier Code", "Dep Airport Code", "Arr Airport Code", "Seats (Total)"}


def log(msg):
    print(time.strftime("%H:%M:%S"), msg, flush=True)


def load_manifest():
    try:
        return json.load(open(MANIFEST))
    except Exception:
        return {}


def save_manifest(m):
    tmp = MANIFEST + ".tmp"
    json.dump(m, open(tmp, "w"), indent=1)
    os.replace(tmp, MANIFEST)


def parse_filename(name):
    """'Europe Jan 2015.xlsx' -> region, period, year."""
    stem = os.path.splitext(name)[0]
    m = re.match(r"(.+?)\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|H1|H2|Dec)?\s*(\d{4})$", stem)
    if not m:
        return stem, "year", None
    region, period, year = m.group(1).strip(), (m.group(2) or "year"), int(m.group(3))
    return region, period, year


def sheet_tables(ws):
    """Yield (header_map, rows_iter) for each header block found in a sheet.
    OAG exports sometimes append several jobs to one workbook; each block starts
    with a row containing the key column names."""
    rows = ws.iter_rows(values_only=True)
    header, idx = None, {}
    batch = []
    for r in rows:
        vals = ["" if v is None else str(v).strip() for v in r]
        if KEY.issubset(set(vals)):
            if header and batch:
                yield idx, batch
            header = vals
            idx = {c: header.index(c) for c in WANTED if c in header}
            batch = []
            continue
        if header and any(v != "" for v in vals):
            batch.append(r)
            if len(batch) >= 200000:
                yield idx, batch
                batch = []
    if header and batch:
        yield idx, batch


def ingest_file(path, region, period, year):
    wb = load_workbook(path, read_only=True, data_only=True)
    cols = {c: [] for c in WANTED}
    n = 0
    for ws in wb.worksheets:
        if ws.title.lower().startswith("note"):
            continue
        for idx, batch in sheet_tables(ws):
            if not idx:
                continue
            for r in batch:
                for c in WANTED:
                    j = idx.get(c)
                    v = r[j] if (j is not None and j < len(r)) else None
                    if c in NUMERIC:
                        try:
                            v = float(v) if v not in (None, "") else None
                        except (TypeError, ValueError):
                            v = None
                    else:
                        v = None if v is None else str(v)
                    cols[c].append(v)
                n += 1
    wb.close()
    if n == 0:
        raise ValueError("no data rows recognised (header names not found)")
    arrays = {c.replace(" ", "_").replace("(", "").replace(")", "").replace("/", "_").lower(): pa.array(cols[c])
              for c in WANTED}
    base = os.path.splitext(os.path.basename(path))[0]
    arrays["source_file"] = pa.array([base] * n)
    arrays["file_region"] = pa.array([region] * n)
    arrays["file_period"] = pa.array([period] * n)
    arrays["file_year"] = pa.array([year] * n, type=pa.int32())
    table = pa.table(arrays)
    os.makedirs(PARQ, exist_ok=True)
    out = os.path.join(PARQ, base + ".parquet")
    pq.write_table(table, out)
    con = duckdb.connect(DB)
    con.execute("CREATE TABLE IF NOT EXISTS schedules AS SELECT * FROM read_parquet(?) LIMIT 0", [out])
    con.execute("DELETE FROM schedules WHERE source_file = ?", [base])
    con.execute("INSERT INTO schedules SELECT * FROM read_parquet(?)", [out])
    con.close()
    return n


def run(one=None, loop=False):
    os.makedirs(RAW, exist_ok=True)
    while True:
        manifest = load_manifest()
        files = sorted(glob.glob(os.path.join(RAW, "*.xlsx")))
        todo = []
        for f in files:
            name = os.path.basename(f)
            if name.startswith("~$"):
                continue
            if one and name != one:
                continue
            rec = manifest.get(name)
            if rec and rec.get("status") == "ok" and rec.get("size") == os.path.getsize(f):
                continue
            todo.append(f)
        if not todo:
            log("nothing new to ingest (%d files complete)" % sum(1 for r in manifest.values() if r.get("status") == "ok"))
        for f in todo:
            name = os.path.basename(f)
            region, period, year = parse_filename(name)
            log("ingesting %s (region=%s period=%s year=%s, %.0f MB)" %
                (name, region, period, year, os.path.getsize(f) / 1e6))
            try:
                n = ingest_file(f, region, period, year)
                manifest[name] = {"status": "ok", "rows": n, "size": os.path.getsize(f),
                                  "ts": time.strftime("%Y-%m-%d %H:%M:%S")}
                log("  ok: %d rows" % n)
            except Exception as e:
                manifest[name] = {"status": "error", "error": str(e)[:300],
                                  "size": os.path.getsize(f),
                                  "ts": time.strftime("%Y-%m-%d %H:%M:%S")}
                log("  ERROR: %s" % e)
            save_manifest(manifest)
        ok = [r for r in manifest.values() if r.get("status") == "ok"]
        log("manifest: %d ok, %d error, %.1fm rows total" %
            (len(ok), sum(1 for r in manifest.values() if r.get("status") == "error"),
             sum(r.get("rows", 0) for r in ok) / 1e6))
        if not loop:
            break
        time.sleep(600)


if __name__ == "__main__":
    one = None
    if "--one" in sys.argv:
        one = sys.argv[sys.argv.index("--one") + 1]
    run(one=one, loop="--loop" in sys.argv)
