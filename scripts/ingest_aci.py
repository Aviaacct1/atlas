"""ACI per-airport traffic ingest (Phase 3 base-year anchor). Author: Avia Solutions.

Reads the decrypted ACI annual Traffic datasets on E: and builds a tidy per-airport
panel: iata, year, region, country_code, city, terminal_pax, domestic, international,
direct_transit, total_pax, movements, total_cargo. Columns are matched by NAME (the
column order shifts between years), the header row is found by locating 'IATA', and the
year-suffixed metric columns are selected per file.

ACI is subscription data: this panel is INTERNAL base-year calibration and history only,
never redistributed in the product (class C, like Sabre).

Data root: E:\\Avia\\Global (bash mount /sessions/.../mnt/Global). Decrypted files live in
data/aci_decrypted/aci_<year>.xlsx.
"""
from __future__ import annotations
import os as _os, sys as _sys; _sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
from avia_forecast.paths import DATA, OEF_DIR, ACI_DIR, ACI_DECRYPT, SABRE_DB, OAG_DB, QSI_REF, PREAGG, QSI_APP, OEF_GDP_XLSX
import json, os, re, sys

import openpyxl

DATA = DATA
DEC = os.path.join(DATA, "aci_decrypted")
MODERN_YEARS = list(range(2013, 2025))     # 2013-2024 clean name-based layout


def _num(v):
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s in ("", "-", "0") or s.upper().startswith("ERROR") or "#" in s:
        return 0.0 if s == "0" else None
    try:
        return float(s)
    except ValueError:
        return None


def _norm(h):
    return re.sub(r"\s+", " ", str(h).strip().lower()) if h is not None else ""


def _pick_sheet(wb, year):
    names = {n.lower(): n for n in wb.sheetnames}
    for cand in (f"annual {year} dataset", f"annual {year}", str(year)):
        if cand in names:
            return names[cand]
    for low, real in names.items():
        if str(year) in low and "dataset" in low:
            return real
    return None


def _find_header(ws, scan=12):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=scan, values_only=True)):
        vals = [_norm(c) for c in row]
        if any(v in ("iata code", "iata", "code") for v in vals):
            return i + 1, vals
    return None, None


def _col_map(headers, year):
    idx = {}
    want = {
        "iata": ("iata code", "iata"),
        "region": ("region",),
        "country_code": ("country code",),
        "country": ("country",),
        "city": ("city",),
        "terminal": (f"passenger terminal {year}",),
        "domestic": (f"passenger domestic {year}",),
        "international": (f"passenger international {year}",),
        "transit": (f"passenger direct transit {year}",),
        "total_pax": (f"passengers {year}",),
        "movements": (f"movements {year}",),
        "cargo": (f"total cargo {year}",),
    }
    for key, names in want.items():
        for j, h in enumerate(headers):
            if h in names:
                idx[key] = j
                break
    return idx


def parse_modern(path, year):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = _pick_sheet(wb, year)
    if sheet is None:
        wb.close(); raise RuntimeError(f"{year}: no data sheet")
    ws = wb[sheet]
    hrow, headers = _find_header(ws)
    if hrow is None:
        wb.close(); raise RuntimeError(f"{year}: no header found in {sheet}")
    idx = _col_map(headers, year)
    for req in ("iata", "terminal"):
        if req not in idx:
            wb.close(); raise RuntimeError(f"{year}: missing column {req} in {sheet}")
    out = []
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        iata = row[idx["iata"]] if idx["iata"] < len(row) else None
        iata = str(iata).strip().upper() if iata is not None else ""
        if not re.fullmatch(r"[A-Z]{3}", iata):
            continue
        def g(k):
            j = idx.get(k)
            return _num(row[j]) if (j is not None and j < len(row)) else None
        term = g("terminal")
        if term is None:
            continue
        out.append({
            "iata": iata, "year": year,
            "region": (row[idx["region"]] if "region" in idx else None),
            "country_code": (row[idx["country_code"]] if "country_code" in idx else None),
            "city": (row[idx["city"]] if "city" in idx else None),
            "terminal_pax": term, "domestic": g("domestic"), "international": g("international"),
            "direct_transit": g("transit"), "total_pax": g("total_pax"),
            "movements": g("movements"), "total_cargo": g("cargo"),
        })
    wb.close()
    return out


def run():
    panel = []
    per_year = {}
    for y in MODERN_YEARS:
        path = os.path.join(DEC, f"aci_{y}.xlsx")
        if not os.path.exists(path):
            print(f"  {y}: file missing, skip"); continue
        rows = parse_modern(path, y)
        panel.extend(rows)
        world = sum(r["terminal_pax"] for r in rows if r["terminal_pax"])
        per_year[y] = (len(rows), world)
        print(f"  {y}: {len(rows):>5} airports, world terminal pax {world/1e9:5.2f}bn")
    os.makedirs(DATA, exist_ok=True)
    json.dump(panel, open(os.path.join(DATA, "aci_panel_2013_2024.json"), "w"))
    print(f"\npanel rows: {len(panel):,}  ->  {DATA}/aci_panel_2013_2024.json")
    return panel, per_year


if __name__ == "__main__":
    run()
