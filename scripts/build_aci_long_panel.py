"""Extend the ACI per-airport panel back in time for elasticity identification.
Adds 1991-2002 (long format) and 2010-2012 (two-row positional files) to the clean
2013-2024 panel, giving ~1991-2024 terminal-passenger history per airport. Country is
backfilled by IATA from the modern panel (+ QSI reference). Author: Avia Solutions.
"""
from __future__ import annotations
import os as _os, sys as _sys; _sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
from avia_forecast.paths import DATA, OEF_DIR, ACI_DIR, ACI_DECRYPT, SABRE_DB, OAG_DB, QSI_REF, PREAGG, QSI_APP, OEF_GDP_XLSX
import csv, json, os, re
import openpyxl

DEC = ACI_DECRYPT
DATA = DATA
QSI_REF = QSI_REF
IATA = re.compile(r"[A-Z]{3}")


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "-") or s.upper().startswith("ERROR") or "#" in s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_1991_2002():
    wb = openpyxl.load_workbook(os.path.join(DEC, "aci_1991_2002.xlsx"), read_only=True, data_only=True)
    ws = wb["ACI 1991-2002"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = str(r[6]).strip().upper() if r[6] else ""
        yr = _num(r[7]); term = _num(r[15])
        if IATA.fullmatch(code) and yr and term:
            out.append((code, int(yr), term))
    wb.close()
    return out


def parse_two_row(fname, headline_year):
    """2011/2012 files: row1 short codes, data from row3. Each carries headline year (Passenger
    Terminal = idx12, Movements idx9) and the prior year (idx27)."""
    wb = openpyxl.load_workbook(os.path.join(DEC, fname), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[-1]] if "ReadMe" in wb.sheetnames[0] else wb[wb.sheetnames[0]]
    # choose the data sheet: the one whose name is the year or has most rows
    ds = None
    for s in wb.sheetnames:
        if str(headline_year) in s or s.lower().startswith("annual"):
            ds = s; break
    ws = wb[ds] if ds else ws
    out = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        code = str(r[4]).strip().upper() if len(r) > 4 and r[4] else ""
        if not IATA.fullmatch(code):
            continue
        cur = _num(r[12]) if len(r) > 12 else None
        prev = _num(r[27]) if len(r) > 27 else None
        if cur:
            out.append((code, headline_year, cur))
        if prev:
            out.append((code, headline_year - 1, prev))
    wb.close()
    return out


def run():
    modern = json.load(open(os.path.join(DATA, "aci_panel_2013_2024.json")))
    iata_country = {}
    for r in modern:
        if r["country_code"] and r["iata"] not in iata_country:
            iata_country[r["iata"]] = r["country_code"]
    for row in csv.DictReader(open(QSI_REF, encoding="utf-8-sig")):
        iata_country.setdefault(row["airport_code"].strip(), row["country_code"].strip())

    # long records: (iata, year, terminal)
    seen = {(r["iata"], r["year"]) for r in modern}
    long = [(r["iata"], r["year"], r["terminal_pax"]) for r in modern if r["terminal_pax"]]
    added = 0
    for rows in (parse_1991_2002(),
                 parse_two_row("aci_2011.xlsx", 2011),
                 parse_two_row("aci_2012.xlsx", 2012)):
        for iata, yr, term in rows:
            if (iata, yr) not in seen:
                seen.add((iata, yr)); long.append((iata, yr, term)); added += 1

    panel = [{"iata": i, "year": y, "terminal_pax": t, "country_code": iata_country.get(i)}
             for i, y, t in long]
    json.dump(panel, open(os.path.join(DATA, "aci_panel_long.json"), "w"))
    yrs = sorted({y for _, y, _ in long})
    from collections import Counter
    c = Counter(y for _, y, _ in long)
    print(f"long panel: {len(panel):,} rows, {added:,} added pre-2013; years {yrs[0]}-{yrs[-1]}")
    print("airports/year:", {y: c[y] for y in yrs})
    return panel


if __name__ == "__main__":
    run()
