"""Ingest an EIU 'World GDP Forecast' vintage xlsx: extract each country's 'Real GDP (% change pa)'
(DGDP) series and cumulate to a forecast GDP index {ISO2:{year:index}} - the GDP a forecaster held at
the vintage date, for the backtest decomposition. Author: Avia Solutions.

Usage: python scripts/ingest_eiu_vintage.py <vintage.xlsx> [base_year] [out.json]
"""
import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from avia_forecast.io_safe import dump_atomic  # atomic write with parse-back (io_safe rule, 23 Aug 2026)


def _rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return list(wb[wb.sheetnames[0]].iter_rows(values_only=True))


def extract_growth(rows):
    """{ISO2: {year: real GDP % change pa}} from the DGDP rows, using the header year columns."""
    def _year(v):
        try:
            iv = int(str(v).strip())
        except (ValueError, TypeError):
            return None
        return iv if 1980 <= iv <= 2060 else None
    ycols = {}
    for r in rows:
        yc = {j: _year(v) for j, v in enumerate(r) if _year(v) is not None}
        if len(yc) >= 8:
            ycols = yc
            break
    if not ycols:
        raise SystemExit("no header year row found")
    growth = {}
    for r in rows:
        title = next((str(c) for c in r[:6] if isinstance(c, str) and "Real GDP (% change pa)" in c), None)
        if not title:
            continue
        code = r[1] if (isinstance(r[1], str) and len(r[1].strip()) == 2) else None
        if not code:
            continue
        series = {ycols[j]: float(r[j]) for j in ycols if isinstance(r[j], (int, float))}
        if series:
            growth[code.strip()] = series
    return growth


def to_index(growth, base_year):
    """Cumulate % change pa into a GDP index anchored at base_year = 1.0."""
    out = {}
    for code, g in growth.items():
        yrs = sorted(y for y in g if y >= base_year)
        if base_year not in g and not yrs:
            continue
        idx, lvl = {}, 1.0
        idx[base_year] = 1.0
        for y in range(base_year + 1, max(yrs) + 1 if yrs else base_year):
            lvl *= (1.0 + g.get(y, 0.0) / 100.0)
            idx[y] = round(lvl, 6)
        out[code] = idx
    return out


def main():
    path = sys.argv[1]
    base_year = int(sys.argv[2]) if len(sys.argv) > 2 else None
    rows = _rows(path)
    growth = extract_growth(rows)
    if base_year is None:
        # infer from filename year, else the earliest fully-forecast year
        base_year = 2015
    idx = to_index(growth, base_year)
    out = sys.argv[3] if len(sys.argv) > 3 else os.path.splitext(path)[0] + f"_gdp_index_{base_year}.json"
    dump_atomic({"base_year": base_year, "gdp_index": idx}, out, indent=0)
    print(f"{len(idx)} countries -> {out}  (base {base_year})")


if __name__ == "__main__":
    main()
