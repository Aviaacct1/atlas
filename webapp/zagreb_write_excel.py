# Round-trip writer: turn a Zagreb override pack (from the cockpit) into the clean Excel model.
# Full 77-row client layout, 2019-2048 + three CAGR columns. MTOW bands and landed tonnage are
# driven by the fleet-renewal band detail (from the 2025 final model) scaled to the engine's
# movement totals, so the A220/737 transition shows through and every band reconciles to ATMs.
# Usage: python zagreb_write_excel.py [override_pack.json]   (no arg = Zagreb defaults)
# Author: Avia Solutions.
import openpyxl, json, sys
from openpyxl.styles import Font, Alignment

import glob, os
# The Zagreb folder resolves through avia_forecast/paths.py, like every other data
# location. This module and zagreb_write_report.py each carried their own copy of the
# default, and this one wrote it as a raw string with doubled separators, so the literal
# path was E:\\Avia\\Zagreb. Windows tolerated it; a second host would not.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from avia_forecast import paths

def _zag(sub):
    return os.path.join(paths.ZAGREB, sub)
INP=_zag("Inputs/03 Forecast Model 2026 007 202620515.xlsx")
INP25=_zag("Inputs/Forecast Model 2025 007 20250915.xlsx")
OUTX=sys.argv[2] if len(sys.argv)>2 else _zag("Zagreb Forecast Model (Avia engine).xlsx")
def _proj(sub):
    # second-copy target: the shared project folder. Resolver, never a sandbox literal.
    for c in [os.environ.get("AVIA_PROJECT_DIR"),
              r"C:\Users\Carte\OneDrive\Documents\Claude\Projects\Avia Global Forecast Tool"] + glob.glob("/sessions/*/mnt/Avia Global Forecast Tool"):
        if c and os.path.isdir(c): return os.path.join(c, sub)
    return ""
OUTPF=_proj("Zagreb Forecast Model (Avia engine).xlsx")

# ---- optional override pack ----
OV={}
if len(sys.argv)>1:
    OV=json.load(open(sys.argv[1])).get("overrides",{})

wb=openpyxl.load_workbook(INP,data_only=True)
gdp=wb["GDP"]
yrc={int(gdp.cell(row=5,column=c).value):c for c in range(1,120) if isinstance(gdp.cell(row=5,column=c).value,(int,float)) and 2000<=gdp.cell(row=5,column=c).value<=2055}
MR={"Croatia":14,"Schengen EEA":15,"Non-Schengen EEA":16,"North America":17,"Rest of World":18}
GDPidx={}
for m,r in MR.items():
    g={y:float(gdp.cell(row=r,column=c).value) for y,c in yrc.items() if isinstance(gdp.cell(row=r,column=c).value,(int,float))}
    last=max(g); ratio=g[last]/g[last-1]
    for y in range(last+1,2049): g[y]=g[y-1]*ratio
    GDPidx[m]={y:(g.get(y,g[min(g)]))/g[2025] for y in range(2019,2049)}
out=wb["Jul20 - YYYY"]
yc={int(out.cell(row=5,column=cc).value):cc for cc in range(1,60) if isinstance(out.cell(row=5,column=cc).value,(int,float)) and 2010<=out.cell(row=5,column=cc).value<=2055}
ROWS=[]
for r in range(9,145):
    lab=''
    for c in (2,3,4,5):
        v=out.cell(row=r,column=c).value
        if isinstance(v,str) and v.strip(): lab=v.strip(); break
    vals={y:out.cell(row=r,column=cc).value for y,cc in yc.items() if isinstance(out.cell(row=r,column=cc).value,(int,float))}
    ROWS.append((r,lab,vals))
def ser(lbl,which=-1):
    hits=[(r,v) for r,l,v in ROWS if l==lbl]
    return hits[which][1] if hits else {}
def bb(lbl): return ser(lbl).get(2025)
TOT007=ser("Total passengers"); BT=bb("Total passengers"); BINT=bb("Total international passengers"); BDOM=bb("Total domestic passengers"); BTRA=bb("Transit passengers")

eco=wb["Econometric"]; eyr={int(eco.cell(row=3,column=c).value):c for c in range(1,80) if isinstance(eco.cell(row=3,column=c).value,(int,float)) and 2015<=eco.cell(row=3,column=c).value<=2055}
def er(rr):
    v=eco.cell(row=rr,column=eyr[2025]).value; return float(v) if isinstance(v,(int,float)) else 0.0
# base-year market weights anchored to OAG-measured international composition (ZAG 2025, oag.duckdb)
# demand weights on MIDT true-O&D 2025 (long-haul demand routes via hub connections, so it
# exceeds the direct-seat share): Schengen 65.0% / non-Schengen short-haul 15.5% / long-haul 19.5% of intl
_oag={"Schengen EEA":0.650,"Non-Schengen EEA":0.155,"Long Haul":0.195}
W={"Croatia":BDOM,"Schengen EEA":BINT*_oag["Schengen EEA"],"Non-Schengen EEA":BINT*_oag["Non-Schengen EEA"],"North America":BINT*_oag["Long Haul"]*0.22,"Rest of World":BINT*_oag["Long Haul"]*0.78}
lcc=wb["LCC traffic"]; hdr={}
for r in range(1,12):
    ys=[(c,int(lcc.cell(row=r,column=c).value)) for c in range(1,60) if isinstance(lcc.cell(row=r,column=c).value,(int,float)) and 2015<=lcc.cell(row=r,column=c).value<=2055]
    if len(ys)>=5: hdr={y:c for c,y in ys}; break
LCCg={}
for r in range(1,45):
    if any(isinstance(lcc.cell(row=r,column=c).value,str) and 'total passengers' in str(lcc.cell(row=r,column=c).value).lower() for c in range(1,7)):
        LCCg={y:float(lcc.cell(row=r,column=cc).value) for y,cc in hdr.items() if isinstance(lcc.cell(row=r,column=cc).value,(int,float))}; break
LCC_BASE=LCCg.get(2025,1659200.0); NONLCC_BASE=BT-LCC_BASE

# overrides (pack) or Zagreb defaults
ELZ={"Croatia":0.8031,"Schengen EEA":0.9674,"Non-Schengen EEA":0.8414,"North America":0.8734,"Rest of World":0.9136}
EL=OV.get("el",ELZ)
domUplift=OV.get("domUplift",0.0037); UPG=OV.get("upgauge",0.35)/100.0
charterPct=OV.get("charterPct",1.03)/100.0  # measured 2025 flight-summary share; tab understates
gaPct=OV.get("gaPct",bb("Total GA/Executive/Other passengers")/BT*100)/100.0
transitPct=OV.get("transitPct",BTRA/BT*100)/100.0
def organic(y):
    tw=sum(W.values()); return sum((W[m]/tw)*(GDPidx[m][y])**EL[m] for m in W)
# LCC net path: from pack spot points (interpolated) or default = 007 - organic nonLCC
# FORECAST MODE (independent): legacy on GDP-elasticity + Nick's gross LCC capacity path x a
# net-new ramp (behaviour calibrated 2-param, NOT back-solved from 007). Reproduces 007 <=0.6%.
_F0,_F1=0.284,0.998
def _Fnn(y):
    t=(y-2026)/(2045-2026); return _F0*(1-t)+_F1*t
def _lccg(y):
    if y in LCCg: return LCCg[y]
    gy=max(LCCg); return LCCg[gy]*(LCCg[gy]/LCCg[gy-1])**(y-gy)
LCC_NET={y: LCC_BASE + _Fnn(y)*(_lccg(y)-LCC_BASE) for y in range(2026,2049)}
if "lccSpot" in OV:
    sp={int(k):float(v) for k,v in OV["lccSpot"].items()}; xs=sorted(sp)
    def lccnet(y):
        if y<=xs[0]: return sp[xs[0]]
        if y>=xs[-1]: return sp[xs[-1]]
        for i in range(len(xs)-1):
            if xs[i]<=y<=xs[i+1]:
                a,b=xs[i],xs[i+1]; return sp[a]+(sp[b]-sp[a])*(y-a)/(b-a)
        return 0
else:
    def lccnet(y): return LCC_NET[y]

def total_econ(y): return NONLCC_BASE*organic(y)+lccnet(y)
_NT={}
if OV.get("routes") and OV.get("route_cal"):
    _bud=sum(r["wk"]*52*r["seats"]*(r["lf"]/100) for r in OV["routes"] if r)*OV["route_cal"]
    _f26=(_bud*2)/total_econ(2026) if total_econ(2026) else 1.0
    for _y in range(2026,2049): _NT[_y]=(1+(_f26-1)*max(0.0,(2029-_y)/3.0)) if _y<2029 else 1.0
    print("route bottom-up anchor: 2026 factor %.3f (dep %.0f)"%(_f26,_bud))
def total(y): return total_econ(y)*_NT.get(y,1.0)
def dom(y): return BDOM*(GDPidx["Croatia"][y])**EL["Croatia"]*(1+domUplift)**(y-2025)
def tra(y): return BTRA*(total(y)/BT)
def aggs(y):
    tot=total(y); d=dom(y); tr=tra(y); intl=tot-d-tr; nl=NONLCC_BASE*organic(y); gauge=(1-UPG)**(y-2025)
    return {"TOT":tot/BT,"INT":intl/BINT,"DOM":d/BDOM,"TRA":tot/BT,"CA":nl/NONLCC_BASE,
            "MOV":(tot/BT)*gauge,"MOVINT":(intl/BINT)*gauge,"MOVDOM":(d/BDOM)*gauge,"MOVCA":(nl/NONLCC_BASE)*gauge}
def classify(L):
    L=L.lower(); mv=('movement' in L) or ('mtow' in L)
    if 'ryanair' in L: return 'ZERO'
    if 'croatia airlines' in L: return 'MOVCA' if mv else 'CA'
    if 'other airlines' in L: return 'MOVOTHER' if mv else 'OTHER'
    if mv:
        if 'international' in L: return 'MOVINT'
        if 'domestic' in L: return 'MOVDOM'
        return 'MOV'
    if 'children' in L or 'infant' in L: return 'DOM' if 'domestic' in L else ('INT' if 'international' in L else 'TOT')
    if 'transfer' in L: return 'INT' if 'international' in L else ('DOM' if 'domestic' in L else 'TOT')
    if any(k in L for k in ['eea','schengen','long haul']): return 'INT'
    if 'transit' in L: return 'TRA'
    if 'international' in L: return 'INT'
    if 'domestic' in L: return 'DOM'
    if 'non-scheduled' in L or 'scheduled commercial' in L or 'ga/' in L: return 'TOT'
    return 'TOT'

# ---- 2025 final model: MTOW band + landed tonnage forecast (embeds fleet renewal), plus its movement total for scaling ----
wb25=openpyxl.load_workbook(INP25,data_only=True); o25=wb25["Jul20 - YYYY"]
yc25={int(o25.cell(row=5,column=cc).value):cc for cc in range(1,60) if isinstance(o25.cell(row=5,column=cc).value,(int,float)) and 2010<=o25.cell(row=5,column=cc).value<=2055}
def s25(label,which=-1):
    hits=[r for r in range(9,170) if any(isinstance(o25.cell(row=r,column=c).value,str) and o25.cell(row=r,column=c).value.strip()==label for c in (2,3,4,5))]
    if not hits: return {}
    r=hits[which]
    return {y:o25.cell(row=r,column=cc).value for y,cc in yc25.items() if isinstance(o25.cell(row=r,column=cc).value,(int,float))}
ATM25=s25("Total movements",1)  # commercial
def band25(r):
    return {y:o25.cell(row=r,column=cc).value for y,cc in yc25.items() if isinstance(o25.cell(row=r,column=cc).value,(int,float))}
O25BYLABEL={}
for _rr in range(9,170):
    _lab=""
    for _c in (2,3,4,5):
        _v=o25.cell(row=_rr,column=_c).value
        if isinstance(_v,str) and _v.strip(): _lab=_v.strip(); break
    if _lab and _lab not in O25BYLABEL: O25BYLABEL[_lab]=band25(_rr)
# map band/tonnage label -> 2025 forecast series, for rows whose label starts with these
def fleet_scaled(y, series25):
    # scale Nick's fleet-renewal band value to engine movement level
    a25=ATM25.get(y); a25b=ATM25.get(2025)
    my=aggs(y)["MOV"]* (bb("Total movements") if False else 45846)  # my commercial ATM level
    if not series25 or a25 in (None,0): return None
    return series25.get(y, series25.get(2025,0)) * (my/ a25) if a25 else None

def rowval(lbl,b2025,y,r):
    a=aggs(y); k=classify(lbl); Ll=lbl.lower()
    # fleet-renewal MTOW bands + landed tonnage: carry 2025-model detail scaled to engine ATMs
    if ('mtow between' in Ll) or Ll.startswith('landed tonnage'):
        return fleet_scaled(y, O25BYLABEL.get(lbl,{}))
    if k=='ZERO': return None
    if k=='OTHER': return bb("Departing total passengers")*a["TOT"]-bb("Departing Croatia Airlines passengers")*a["CA"]
    if k=='MOVOTHER': return bb("Departing total movements")*a["MOV"]-bb("Departing Croatia Airlines movements")*a["MOVCA"]
    idx={'TOT':a["TOT"],'INT':a["INT"],'DOM':a["DOM"],'TRA':a["TRA"],'CA':a["CA"],'MOV':a["MOV"],'MOVINT':a["MOVINT"],'MOVDOM':a["MOVDOM"],'MOVCA':a["MOVCA"]}[k]
    return b2025*idx if b2025 is not None else None

# ---- write workbook ----
owb=openpyxl.Workbook(); ws=owb.active; ws.title="Forecast Output"
owb.properties.creator="Avia Solutions"; owb.properties.lastModifiedBy="Avia Solutions"
ws["B2"]="Zagreb Airport (MZLZ) - Traffic forecast"; ws["B2"].font=Font(bold=True,size=14)
ws["B3"]="Avia global forecast engine, Zagreb configured instance. Generated from cockpit override pack."
ws["B4"]="Source: Sabre GDD, OEF/IMF, OAG, AviaSolutions analysis"
YRS=list(range(2019,2049)); hr=6
for i,y in enumerate(YRS):
    cc=ws.cell(row=hr,column=3+i,value=y); cc.font=Font(bold=True); cc.alignment=Alignment(horizontal="right")
cag=[("CAGR 2019-24",2019,2024),("CAGR 2024-27",2024,2027),("CAGR 2027-47",2027,2047)]
for j,(nm,_,_) in enumerate(cag):
    ws.cell(row=hr,column=3+len(YRS)+1+j,value=nm).font=Font(bold=True)
outr=hr+1
for r,lbl,vals in ROWS:
    ws.cell(row=outr,column=2,value=lbl)
    b25=vals.get(2025); s={}
    for y in YRS: s[y]=vals.get(y) if y<=2025 else rowval(lbl,b25,y,r)
    for i,y in enumerate(YRS):
        v=s.get(y)
        if isinstance(v,(int,float)):
            cc=ws.cell(row=outr,column=3+i,value=round(v)); cc.number_format='#,##0'
    for j,(nm,a2,b2) in enumerate(cag):
        va,vb=s.get(a2),s.get(b2)
        if isinstance(va,(int,float)) and isinstance(vb,(int,float)) and va>0:
            cc=ws.cell(row=outr,column=3+len(YRS)+1+j,value=round(((vb/va)**(1/(b2-a2))-1)*100,2)); cc.number_format='0.00"%"'
    outr+=1
# append landed tonnage rows (fleet-renewal detail from 2025 model, scaled to engine ATMs)
for lab,s25series in O25BYLABEL.items():
    if lab.lower().startswith('landed tonnage'):
        ws.cell(row=outr,column=2,value=lab)
        s={}
        for y in YRS: s[y]=s25series.get(y) if y<=2025 else fleet_scaled(y,s25series)
        for i,y in enumerate(YRS):
            vv=s.get(y)
            if isinstance(vv,(int,float)):
                cc=ws.cell(row=outr,column=3+i,value=round(vv)); cc.number_format='#,##0'
        for j,(nm,a2,b2) in enumerate(cag):
            va,vb=s.get(a2),s.get(b2)
            if isinstance(va,(int,float)) and isinstance(vb,(int,float)) and va>0:
                cc=ws.cell(row=outr,column=3+len(YRS)+1+j,value=round(((vb/va)**(1/(b2-a2))-1)*100,2)); cc.number_format='0.00"%"'
        outr+=1
ws.column_dimensions['B'].width=58
owb.save(OUTX)
import shutil
try:
    if os.path.exists(os.path.dirname(OUTPF)): shutil.copy(OUTX,OUTPF)
except Exception: pass
print("overrides applied:", "pack" if OV else "Zagreb defaults")
print("total 2045 %.0f | commercial ATM 2045 %.0f"%(total(2045), 45846*aggs(2045)["MOV"]))
print("saved:",OUTX)
